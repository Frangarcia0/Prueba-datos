import subprocess, sys
for pkg in ["requests", "openpyxl", "python-docx", "pandas"]:
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--break-system-packages", "-q"])

import argparse
import io
import logging
import os
import re
import urllib.parse
from pathlib import Path

import pandas as pd
import requests
from docx import Document

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    filename="etl_hestia.log",
    level=logging.WARNING,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("etl_hestia")


# ── file source abstraction ───────────────────────────────────────────────────
class LocalSource:
    def __init__(self, root: Path):
        self.root = root

    def list_dir(self, rel_path: str) -> list[dict]:
        p = self.root / rel_path
        if not p.exists():
            return []
        entries = []
        for child in sorted(p.iterdir()):
            entries.append({
                "type": "dir" if child.is_dir() else "file",
                "name": child.name,
                "path": str(child.relative_to(self.root)),
            })
        return entries

    def read_file(self, rel_path: str) -> bytes:
        return (self.root / rel_path).read_bytes()

    def walk_files(self, rel_path: str, ext: str) -> list[dict]:
        results = []
        for item in self.list_dir(rel_path):
            if item["type"] == "file" and item["name"].lower().endswith(ext):
                results.append(item)
            elif item["type"] == "dir":
                results.extend(self.walk_files(item["path"], ext))
        return results


class GitHubSource:
    API = "https://api.github.com/repos/{repo}/contents/{path}"
    RAW = "https://raw.githubusercontent.com/{repo}/{ref}/{path}"

    def __init__(self, repo: str, ref: str = "main", token: str = ""):
        self.repo = repo
        self.ref = ref
        self.headers = {"Authorization": f"token {token}"} if token else {}

    def list_dir(self, rel_path: str) -> list[dict]:
        url = self.API.format(repo=self.repo, path=urllib.parse.quote(rel_path, safe="/"))
        r = requests.get(url, params={"ref": self.ref}, headers=self.headers, timeout=30)
        r.raise_for_status()
        return [{"type": i["type"], "name": i["name"], "path": i["path"]} for i in r.json()]

    def read_file(self, rel_path: str) -> bytes:
        url = self.RAW.format(
            repo=self.repo, ref=self.ref,
            path=urllib.parse.quote(rel_path, safe="/"),
        )
        r = requests.get(url, headers=self.headers, timeout=60)
        r.raise_for_status()
        return r.content

    def walk_files(self, rel_path: str, ext: str) -> list[dict]:
        results = []
        try:
            items = self.list_dir(rel_path)
        except Exception as e:
            log.warning("Cannot list %s: %s", rel_path, e)
            return results
        for item in items:
            if item["type"] == "file" and item["name"].lower().endswith(ext):
                results.append(item)
            elif item["type"] == "dir":
                results.extend(self.walk_files(item["path"], ext))
        return results


# ── utility ───────────────────────────────────────────────────────────────────
def sala_fmt(val) -> str:
    if val is None:
        return ""
    try:
        n = int(float(str(val)))
        return f"SB-{n:03d}"
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s != "None" else ""


def extract_section_code(text: str) -> str:
    codes = re.findall(r"[A-Z0-9]{3,5}", str(text).upper())
    return codes[0] if codes else str(text).strip()


SECTION_HEADER_RE = re.compile(
    r"(\w+)\s+[Ss]ecci[oó]n\s+(\w+)\s+(\d{1,2}:\d{2})-(\d{1,2}:\d{2})",
    re.IGNORECASE | re.UNICODE,
)


def parse_section_header(header: str):
    """Return (dia, seccion, hora_inicio, hora_fin) or None."""
    m = SECTION_HEADER_RE.search(str(header))
    if m:
        return m.group(1).lower(), m.group(2), m.group(3), m.group(4)
    return None


# ── TIPO 1 — PLANIFICACIÓN ────────────────────────────────────────────────────
def process_planificacion_sheet(ws, sheet_name: str, semestre: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    data_rows = rows[1:]

    section_cols = []  # (col_idx, dia, seccion, hora_inicio, hora_fin)
    col_map = {}       # UPPERCASE_KEY → col_idx

    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        parsed = parse_section_header(str(cell))
        if parsed:
            section_cols.append((i, *parsed))
        else:
            key = str(cell).strip().upper()
            col_map[key] = i

    def find_col(*keys):
        for k in keys:
            for map_key in col_map:
                if k in map_key:
                    return col_map[map_key]
        return None

    idx_sala = find_col("SALA")
    idx_docente = find_col("DOCENTE")
    idx_horario = find_col("HORARIO")
    idx_seccion = find_col("SECCI")

    codigo_asig = re.sub(r"\s+", "", sheet_name).upper()

    records = []
    for row in data_rows:
        if all(v is None for v in row):
            continue

        sala = sala_fmt(row[idx_sala] if idx_sala is not None and idx_sala < len(row) else None)
        docente = ""
        if idx_docente is not None and idx_docente < len(row):
            docente = str(row[idx_docente]).strip() if row[idx_docente] else ""

        if section_cols:
            for col_i, dia, sec, h_ini, h_fin in section_cols:
                val = row[col_i] if col_i < len(row) else None
                if val is None or str(val).strip() in ("", "None"):
                    continue
                records.append({
                    "email_docente": "",
                    "codigo_asignatura": codigo_asig,
                    "seccion": sec,
                    "semestre": semestre,
                    "sala": sala,
                    "dia_semana": dia,
                    "hora_inicio": h_ini,
                    "hora_fin": h_fin,
                })
        else:
            horario = row[idx_horario] if idx_horario is not None and idx_horario < len(row) else None
            seccion = row[idx_seccion] if idx_seccion is not None and idx_seccion < len(row) else None
            sec_str = str(seccion).strip() if seccion else ""
            if not sec_str or sec_str == "None":
                continue

            dia, h_ini, h_fin = "", "", ""
            if horario:
                m = re.search(r"(\w+)\s+(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", str(horario))
                if m:
                    dia, h_ini, h_fin = m.group(1).lower(), m.group(2), m.group(3)

            records.append({
                "email_docente": "",
                "codigo_asignatura": codigo_asig,
                "seccion": extract_section_code(sec_str),
                "semestre": semestre,
                "sala": sala,
                "dia_semana": dia,
                "hora_inicio": h_ini,
                "hora_fin": h_fin,
            })

    return records


def process_planificacion_xlsx(content: bytes, semestre: str, filename: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            recs = process_planificacion_sheet(ws, sheet_name, semestre)
            records.extend(recs)
            if not recs:
                log.warning("Hoja sin datos útiles: %s / %s", filename, sheet_name)
        except Exception as e:
            log.warning("Error en hoja %s / %s: %s", filename, sheet_name, e)
    return records


# ── TIPO 2 — INSUMOS ─────────────────────────────────────────────────────────
SALA_MAP = {
    "BANCO SANGRE": "Banco de Sangre",
    "BANCO DE SANGRE": "Banco de Sangre",
    "ODONTOLOGIA": "Odontología",
    "ODONTOLOGÍA": "Odontología",
    "QUIMICA": "Química",
    "QUÍMICA": "Química",
    "TENS": "TENS",
    "FARMACIA": "Farmacia",
}


def folder_to_sala(folder_name: str) -> str:
    up = folder_name.upper()
    for key, val in SALA_MAP.items():
        if key in up:
            return val
    return folder_name.title()


def cell_val(ws, row: int, col: int):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def process_insumos_sheet(ws, sheet_name: str, sala: str) -> list[dict]:
    max_row = ws.max_row
    categoria = ""

    for r in range(1, min(20, max_row + 1)):
        for c in range(1, 8):
            v = cell_val(ws, r, c)
            if v is None:
                continue
            s = str(v).strip()
            if "Nombre taller" in s or "NOMBRE TALLER" in s.upper():
                cat_val = (cell_val(ws, r, c + 3)
                           or cell_val(ws, r, c + 2)
                           or cell_val(ws, r, c + 1))
                if cat_val:
                    categoria = str(cat_val).strip()

    # find header row: row after "INDICAR CANTIDAD"
    header_row_idx = None
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = cell_val(ws, r, c)
            if v and "INDICAR CANTIDAD" in str(v).upper():
                header_row_idx = r + 1
                break
        if header_row_idx:
            break

    if not header_row_idx or header_row_idx > max_row:
        log.warning("No se encontró tabla de insumos en hoja: %s", sheet_name)
        return []

    records = []
    SKIP_NAMES = {"", "None", "Insumo", "INSUMO", "Equipos", "EQUIPOS",
                  "Nombre", "NOMBRE", "N/A", "-"}

    for r in range(header_row_idx + 1, max_row + 1):
        # left: B(2) C(3) D(4) E(5) F(6)
        insumo_nombre = cell_val(ws, r, 2)
        cant_taller   = cell_val(ws, r, 4)
        cant_semestre = cell_val(ws, r, 5)
        semana_l      = cell_val(ws, r, 6)

        name_str = str(insumo_nombre).strip() if insumo_nombre else ""
        if name_str and name_str not in SKIP_NAMES:
            stock = cant_semestre if cant_semestre is not None else cant_taller
            records.append({
                "nombre": name_str,
                "tipo": "insumo",
                "sku": "",
                "codigo_barras": "",
                "descripcion": f"Cant. por taller: {cant_taller or ''} | Semana: {semana_l or ''}",
                "stock_actual": stock if stock is not None else "",
                "stock_minimo": "",
                "costo_unitario": "",
                "sala": sala,
                "categoria": categoria,
            })

        # right: H(8) I(9) J(10) K(11) L(12)
        equipo_nombre = cell_val(ws, r, 8)
        cant_eq_sem   = cell_val(ws, r, 9)
        semana_r      = cell_val(ws, r, 12)

        eq_str = str(equipo_nombre).strip() if equipo_nombre else ""
        if eq_str and eq_str not in SKIP_NAMES:
            records.append({
                "nombre": eq_str,
                "tipo": "implemento",
                "sku": "",
                "codigo_barras": "",
                "descripcion": f"Cant. por semestre: {cant_eq_sem or ''} | Semana: {semana_r or ''}",
                "stock_actual": cant_eq_sem if cant_eq_sem is not None else "",
                "stock_minimo": "",
                "costo_unitario": "",
                "sala": sala,
                "categoria": categoria,
            })

    return records


def process_insumos_xlsx(content: bytes, folder_name: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sala = folder_to_sala(folder_name)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            recs = process_insumos_sheet(ws, sheet_name, sala)
            records.extend(recs)
            if not recs:
                log.warning("Hoja insumos sin datos: %s / %s", folder_name, sheet_name)
        except Exception as e:
            log.warning("Error en hoja insumos %s / %s: %s", folder_name, sheet_name, e)
    return records


# ── TIPO 3 — GUÍAS DE TALLER ─────────────────────────────────────────────────
def process_guia_docx(content: bytes, asig_code: str) -> list[dict]:
    doc = Document(io.BytesIO(content))
    records = []
    for table in doc.tables:
        taller_nombre = None
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells]
            if len(cells) >= 2 and "Nombre del Taller" in cells[0]:
                taller_nombre = cells[1] if cells[1] else (cells[2] if len(cells) > 2 else "")
                break
        if taller_nombre and taller_nombre.strip():
            records.append({
                "Nombre de Taller": taller_nombre.strip(),
                "Fecha": "",
                "Sala": "",
                "Horario": "",
                "Docente": "",
                "Seccion": "",
            })
    return records


# ── orchestration ─────────────────────────────────────────────────────────────
def find_subdir(src, base: str, keyword: str) -> str | None:
    """Find a sub-directory whose name contains keyword (case-insensitive)."""
    try:
        items = src.list_dir(base)
    except Exception:
        return None
    for item in items:
        if item["type"] == "dir" and keyword.upper() in item["name"].upper():
            return item["path"]
    return None


def collect_planificacion(src, semestre_filter) -> tuple[dict, dict, dict]:
    sems = [1, 2] if semestre_filter is None else [semestre_filter]
    all_recs = {1: [], 2: []}
    file_counts = {1: 0, 2: 0}
    warn_counts = {1: 0, 2: 0}

    for s in sems:
        base = f"Semestre_{s}"
        plan_dir = find_subdir(src, base, "PLANIFICACION")
        if not plan_dir:
            log.warning("No se encontró carpeta PLANIFICACION en Semestre_%s", s)
            continue

        files = src.walk_files(plan_dir, ".xlsx")
        semestre_str = f"2026-{s}"

        for f in files:
            try:
                content = src.read_file(f["path"])
                recs = process_planificacion_xlsx(content, semestre_str, f["name"])
                all_recs[s].extend(recs)
                file_counts[s] += 1
                if not recs:
                    warn_counts[s] += 1
            except Exception as e:
                log.warning("Error planificacion %s: %s", f["path"], e)
                warn_counts[s] += 1

    return all_recs, file_counts, warn_counts


def collect_insumos(src, semestre_filter) -> tuple[dict, dict, dict]:
    sems = [1, 2] if semestre_filter is None else [semestre_filter]
    all_recs = {1: [], 2: []}
    file_counts = {1: 0, 2: 0}
    warn_counts = {1: 0, 2: 0}

    for s in sems:
        base = f"Semestre_{s}"
        ins_dir = find_subdir(src, base, "INSUMOS")
        if not ins_dir:
            log.warning("No se encontró carpeta INSUMOS en Semestre_%s", s)
            continue

        try:
            sub_dirs = [i for i in src.list_dir(ins_dir) if i["type"] == "dir"]
        except Exception as e:
            log.warning("No se puede listar insumos sub-dirs S%s: %s", s, e)
            continue

        for sub in sub_dirs:
            folder_name = sub["name"]
            files = src.walk_files(sub["path"], ".xlsx")
            for f in files:
                try:
                    content = src.read_file(f["path"])
                    recs = process_insumos_xlsx(content, folder_name)
                    all_recs[s].extend(recs)
                    file_counts[s] += 1
                    if not recs:
                        warn_counts[s] += 1
                except Exception as e:
                    log.warning("Error insumos %s: %s", f["path"], e)
                    warn_counts[s] += 1

    return all_recs, file_counts, warn_counts


def collect_guias(src, semestre_filter) -> tuple[dict, dict, dict]:
    sems = [1, 2] if semestre_filter is None else [semestre_filter]
    all_recs = {1: [], 2: []}
    file_counts = {1: 0, 2: 0}
    warn_counts = {1: 0, 2: 0}

    for s in sems:
        base = f"Semestre_{s}"
        guias_dir = find_subdir(src, base, "GUIAS")
        if not guias_dir:
            log.warning("No se encontró carpeta GUIAS en Semestre_%s", s)
            continue

        try:
            disc_dirs = [i for i in src.list_dir(guias_dir) if i["type"] == "dir"]
        except Exception as e:
            log.warning("No se puede listar guias S%s: %s", s, e)
            continue

        for disc in disc_dirs:
            try:
                asig_dirs = [i for i in src.list_dir(disc["path"]) if i["type"] == "dir"]
            except Exception as e:
                log.warning("No se puede listar asig dirs %s: %s", disc["path"], e)
                continue

            for asig in asig_dirs:
                asig_code = re.sub(r"\s+", "", asig["name"]).upper()
                files = src.walk_files(asig["path"], ".docx")
                for f in files:
                    try:
                        content = src.read_file(f["path"])
                        recs = process_guia_docx(content, asig_code)
                        all_recs[s].extend(recs)
                        file_counts[s] += 1
                        if not recs:
                            warn_counts[s] += 1
                            log.warning("Sin talleres extraídos: %s", f["path"])
                    except Exception as e:
                        log.warning("Error guia %s: %s", f["path"], e)
                        warn_counts[s] += 1

    return all_recs, file_counts, warn_counts


# ── output ────────────────────────────────────────────────────────────────────
PLAN_COLS = ["email_docente", "codigo_asignatura", "seccion", "semestre",
             "sala", "dia_semana", "hora_inicio", "hora_fin"]
INSUMOS_COLS = ["nombre", "tipo", "sku", "codigo_barras", "descripcion",
                "stock_actual", "stock_minimo", "costo_unitario", "sala", "categoria"]
GUIAS_COLS = ["Nombre de Taller", "Fecha", "Sala", "Horario", "Docente", "Seccion"]


def save_csv(records: list, cols: list, path: str) -> int:
    df = pd.DataFrame(records, columns=cols) if records else pd.DataFrame(columns=cols)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return len(df)


def print_summary_line(label: str, rows: int, files: int, warns: int):
    icon = "⚠️ " if warns > 0 else "✅"
    tail = f" ({warns} con advertencias, ver etl_hestia.log)" if warns > 0 else ""
    print(f"{icon} {label} — {rows} filas generadas desde {files} archivos{tail}")


def validate_csv(path: str, label: str):
    if not os.path.exists(path):
        return
    df = pd.read_csv(path, nrows=3)
    if df.empty:
        print(f"\n   [{label}] — sin filas")
        return
    print(f"\n   Primeras filas de {label}:")
    print(df.to_string(index=False))


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="ETL Hestia — DuocUC San Bernardo")
    parser.add_argument("--repo", required=True, help="usuario/repo de GitHub")
    parser.add_argument("--tipo", required=True,
                        choices=["all", "planificacion", "insumos", "guias"])
    parser.add_argument("--semestre", type=int, choices=[1, 2], default=None)
    parser.add_argument("--local-path", default=None,
                        help="Ruta local del repo clonado (evita rate limit de GitHub)")
    parser.add_argument("--token", default=os.environ.get("GITHUB_TOKEN", ""),
                        help="Token de GitHub (o env GITHUB_TOKEN)")
    parser.add_argument("--ref", default="main", help="branch/tag/SHA de GitHub")
    args = parser.parse_args()

    # choose source
    if args.local_path:
        src = LocalSource(Path(args.local_path))
        print(f"📂 Leyendo archivos locales desde: {args.local_path}")
    else:
        # auto-detect: if cwd looks like the repo root, use local
        cwd = Path.cwd()
        if (cwd / "Semestre_1").exists() or (cwd / "Semestre_2").exists():
            src = LocalSource(cwd)
            print(f"📂 Repositorio detectado localmente en: {cwd}")
        else:
            src = GitHubSource(args.repo, args.ref, args.token)
            print(f"🌐 Descargando desde GitHub: {args.repo} @ {args.ref}")

    os.makedirs("output", exist_ok=True)
    sem_filter = args.semestre
    summary = []

    if args.tipo in ("all", "planificacion"):
        print("⏳ Procesando planificaciones...")
        recs, fcount, wcount = collect_planificacion(src, sem_filter)
        for s in ([sem_filter] if sem_filter else [1, 2]):
            fname = f"output/horario_academico_S{s}.csv"
            n = save_csv(recs[s], PLAN_COLS, fname)
            print_summary_line(f"horario_academico_S{s}.csv", n, fcount[s], wcount[s])
            summary.append((fname, f"horario_academico_S{s}.csv"))

    if args.tipo in ("all", "insumos"):
        print("⏳ Procesando insumos...")
        recs, fcount, wcount = collect_insumos(src, sem_filter)
        for s in ([sem_filter] if sem_filter else [1, 2]):
            fname = f"output/insumos_S{s}.csv"
            n = save_csv(recs[s], INSUMOS_COLS, fname)
            print_summary_line(f"insumos_S{s}.csv", n, fcount[s], wcount[s])
            summary.append((fname, f"insumos_S{s}.csv"))

    if args.tipo in ("all", "guias"):
        print("⏳ Procesando guías de taller...")
        recs, fcount, wcount = collect_guias(src, sem_filter)
        for s in ([sem_filter] if sem_filter else [1, 2]):
            fname = f"output/programacion_talleres_S{s}.csv"
            n = save_csv(recs[s], GUIAS_COLS, fname)
            print_summary_line(f"programacion_talleres_S{s}.csv", n, fcount[s], wcount[s])
            summary.append((fname, f"programacion_talleres_S{s}.csv"))

    print("\n── Validación (primeras 3 filas por CSV) ────────────────────────────")
    for fpath, flabel in summary:
        validate_csv(fpath, flabel)

    print("\n✅ ETL completado. Logs detallados en etl_hestia.log")


if __name__ == "__main__":
    main()
