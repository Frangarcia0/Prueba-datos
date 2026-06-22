import subprocess, sys
for pkg in ["openpyxl", "pandas"]:
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--break-system-packages", "-q"])

import os
import re
import logging
from pathlib import Path

import openpyxl
import pandas as pd

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    filename="catalogo.log",
    level=logging.WARNING,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("catalogo")

ROOT = Path(__file__).parent
OUTPUT_CSV = ROOT / "output" / "catalogo_asignaturas.csv"
COLS = ["codigo_asignatura", "nombre_modulo", "carrera", "jornada", "semestre", "fuente"]

# ── normalizers ───────────────────────────────────────────────────────────────
CODE_RE = re.compile(r"[A-Z]{2,4}\s?\d{3,4}", re.IGNORECASE)


def norm_code(raw: str) -> str:
    return re.sub(r"\s+", "", raw).upper().strip()


def extract_code_from_text(text: str) -> str:
    """Return first asig code found in text, normalised."""
    m = CODE_RE.search(str(text))
    return norm_code(m.group()) if m else ""


CARRERA_KEYWORDS = [
    ("BANCO DE SANGRE", "Banco de Sangre"),
    ("BANCO SANGRE",    "Banco de Sangre"),
    ("TSLB",           "Banco de Sangre"),
    ("ENFERMERIA",     "Enfermería"),
    ("ENFERMERÍA",     "Enfermería"),
    ("ODONTOLOGIA",    "Odontología"),
    ("ODONTOLOGÍA",    "Odontología"),
    ("QUIMICA Y FARMACIA", "Química y Farmacia"),
    ("QUÍMICA Y FARMACIA", "Química y Farmacia"),
    ("QUIMICA",        "Química y Farmacia"),
    ("QUÍMICA",        "Química y Farmacia"),
    ("FARMACIA",       "Química y Farmacia"),
    ("TENS",           "TENS"),
]

CARRERA_FROM_RAW = {
    "técnico en laboratorio clínico y banco de sangre": "Banco de Sangre",
    "tecnico en laboratorio clinico y banco de sangre": "Banco de Sangre",
    "técnico en enfermería":  "Enfermería",
    "tecnico en enfermeria":  "Enfermería",
    "tens":                   "TENS",
    "técnico en química y farmacia": "Química y Farmacia",
    "tecnico en quimica y farmacia": "Química y Farmacia",
}


def infer_carrera(text: str) -> str:
    low = str(text).strip().lower()
    # exact match on known raw carrera strings
    if low in CARRERA_FROM_RAW:
        return CARRERA_FROM_RAW[low]
    # keyword search in upper
    up = str(text).upper()
    for kw, val in CARRERA_KEYWORDS:
        if kw in up:
            return val
    return ""


def infer_jornada(text: str) -> str:
    up = str(text).upper()
    if "VESPERTINO" in up:
        return "Vespertino"
    if "DIURNO" in up:
        return "Diurno"
    return ""


def infer_semestre(path: Path) -> str:
    for part in path.parts:
        if part == "Semestre_1":
            return "2026-1"
        if part == "Semestre_2":
            return "2026-2"
    return ""


def cell(ws, row: int, col: int):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


# ── Source 1: insumos xlsx ────────────────────────────────────────────────────
def scan_insumos_xlsx(path: Path) -> list[dict]:
    """Extract one record per sheet from an insumos xlsx."""
    semestre = infer_semestre(path)
    # jornada from full path
    jornada = infer_jornada(str(path))

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        log.warning("No se pudo abrir insumos xlsx %s: %s", path, e)
        return []

    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        carrera_raw = ""
        nombre_taller = ""
        sigla = ""

        for r in range(1, 25):
            label_b = str(cell(ws, r, 2) or "").strip()
            val_d   = str(cell(ws, r, 4) or "").strip()
            val_c   = str(cell(ws, r, 3) or "").strip()

            # Some files put value in col C instead of D
            val = val_d if val_d else val_c

            if "Carrera" in label_b and val:
                carrera_raw = val
            elif "Nombre taller" in label_b and val:
                nombre_taller = val
            elif "Sigla" in label_b and val:
                sigla = val
            # stop scanning after INDICAR CANTIDAD header row
            if "INDICAR CANTIDAD" in str(cell(ws, r, 2) or "") + str(cell(ws, r, 3) or ""):
                break

        if not nombre_taller and not sigla:
            log.warning("Hoja sin metadatos en %s / %s", path.name, sheet_name)
            continue

        codigo = norm_code(sigla) if sigla else extract_code_from_text(nombre_taller)
        carrera = infer_carrera(carrera_raw) if carrera_raw else infer_carrera(str(path))

        records.append({
            "codigo_asignatura": codigo,
            "nombre_modulo": nombre_taller.strip() or sheet_name.strip(),
            "carrera": carrera,
            "jornada": jornada,
            "semestre": semestre,
            "fuente": "insumos_xlsx",
        })

    return records


def collect_from_insumos(root: Path) -> list[dict]:
    records = []
    for sem_dir in sorted(root.glob("Semestre_*")):
        for insumos_dir in sem_dir.iterdir():
            if not insumos_dir.is_dir() or "INSUMOS" not in insumos_dir.name.upper():
                continue
            for xlsx in insumos_dir.rglob("*.xlsx"):
                try:
                    records.extend(scan_insumos_xlsx(xlsx))
                except Exception as e:
                    log.warning("Error en insumos xlsx %s: %s", xlsx, e)
    return records


# ── Source 2: guías folder names ──────────────────────────────────────────────
def parse_asig_folder(folder_name: str) -> tuple[str, str]:
    """Return (codigo, nombre_modulo) from a guías asignatura folder name."""
    # Strategy: find code pattern anywhere in name; the rest is nombre
    raw = folder_name.strip()

    # Patterns like "NOMBRE - CODE" or "NOMBRE OK CODE"
    m_sep = re.search(r"[-–]\s*([A-Z]{2,4}\s?\d{3,4})\s*$", raw, re.IGNORECASE)
    if m_sep:
        code = norm_code(m_sep.group(1))
        nombre = raw[:m_sep.start()].strip().rstrip("-–").strip()
        return code, nombre

    m_ok = re.search(r"\bOK\s+([A-Z]{2,4}\s?\d{3,4})\s*$", raw, re.IGNORECASE)
    if m_ok:
        code = norm_code(m_ok.group(1))
        nombre = raw[:m_ok.start()].strip()
        return code, nombre

    # Code is the entire folder name (e.g. "APS 1121", "ATS1111")
    if CODE_RE.fullmatch(raw.replace(" ", "")) or CODE_RE.fullmatch(raw):
        return norm_code(raw), ""

    # Code appears somewhere but folder name is mostly text
    m = CODE_RE.search(raw)
    if m:
        code = norm_code(m.group())
        nombre = re.sub(re.escape(m.group()), "", raw).strip(" -_")
        return code, nombre

    # No code found → folder is module name only
    return "", raw


DISC_CARRERA = {
    "BANCO SANGRE": "Banco de Sangre",
    "BANCO DE SANGRE": "Banco de Sangre",
    "ODONTOLOGIA": "Odontología",
    "ODONTOLOGÍA": "Odontología",
    "TENS": "TENS",
}


def disc_to_carrera(disc_folder_name: str) -> str:
    up = disc_folder_name.upper()
    for k, v in DISC_CARRERA.items():
        if k in up:
            return v
    return infer_carrera(disc_folder_name)


def collect_from_guias(root: Path) -> list[dict]:
    records = []
    for sem_dir in sorted(root.glob("Semestre_*")):
        semestre = "2026-1" if "1" in sem_dir.name else "2026-2"
        for guias_top in sem_dir.iterdir():
            if not guias_top.is_dir() or "GUIAS" not in guias_top.name.upper():
                continue
            # depth: guias_top / discipline / asig_folder
            for disc_dir in sorted(guias_top.iterdir()):
                if not disc_dir.is_dir():
                    continue
                carrera = disc_to_carrera(disc_dir.name)
                jornada = infer_jornada(str(disc_dir))

                for asig_dir in sorted(disc_dir.iterdir()):
                    if not asig_dir.is_dir():
                        continue
                    codigo, nombre = parse_asig_folder(asig_dir.name)
                    records.append({
                        "codigo_asignatura": codigo,
                        "nombre_modulo": nombre or asig_dir.name.strip(),
                        "carrera": carrera,
                        "jornada": jornada,
                        "semestre": semestre,
                        "fuente": "carpeta_guias",
                    })

    return records


# ── Source 3: planificación xlsx ──────────────────────────────────────────────
def scan_planificacion_xlsx(path: Path) -> list[dict]:
    semestre = infer_semestre(path)
    jornada  = infer_jornada(path.name)
    carrera  = infer_carrera(path.name)
    # fallback carrera from parent folder name
    if not carrera:
        carrera = infer_carrera(path.parent.name)

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        log.warning("No se pudo abrir planificacion xlsx %s: %s", path, e)
        return []

    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # sheet name itself may be the code (e.g. TSLB file: BSS2111)
        sheet_clean = sheet_name.strip()
        if CODE_RE.fullmatch(sheet_clean.replace(" ", "")):
            codigo = norm_code(sheet_clean)
            nombre = ""
        else:
            # search first 8 rows for a code
            codigo = ""
            for r in range(1, 9):
                for c in range(1, 12):
                    v = str(cell(ws, r, c) or "")
                    found = extract_code_from_text(v)
                    if found:
                        codigo = found
                        break
                if codigo:
                    break
            nombre = sheet_clean

        records.append({
            "codigo_asignatura": codigo,
            "nombre_modulo": nombre,
            "carrera": carrera,
            "jornada": jornada,
            "semestre": semestre,
            "fuente": "planificacion_xlsx",
        })

    return records


def collect_from_planificacion(root: Path) -> list[dict]:
    records = []
    for sem_dir in sorted(root.glob("Semestre_*")):
        for plan_dir in sem_dir.iterdir():
            if not plan_dir.is_dir() or "PLANIFICACION" not in plan_dir.name.upper():
                continue
            for xlsx in sorted(plan_dir.glob("*.xlsx")):
                try:
                    records.extend(scan_planificacion_xlsx(xlsx))
                except Exception as e:
                    log.warning("Error en planificacion xlsx %s: %s", xlsx, e)
    return records


# ── deduplication & merge ─────────────────────────────────────────────────────
SOURCE_PRIORITY = {"insumos_xlsx": 0, "carpeta_guias": 1, "planificacion_xlsx": 2}


def deduplicate(records: list[dict]) -> list[dict]:
    """
    Keep the highest-priority record for each (codigo, semestre) pair.
    Records without a code are kept as-is (they represent modules
    whose codes couldn't be determined).
    """
    # separate coded vs uncoded
    coded   = [r for r in records if r["codigo_asignatura"]]
    uncoded = [r for r in records if not r["codigo_asignatura"]]

    best: dict[tuple, dict] = {}
    for rec in coded:
        key = (rec["codigo_asignatura"], rec["semestre"])
        existing = best.get(key)
        if existing is None:
            best[key] = rec
        else:
            p_new = SOURCE_PRIORITY.get(rec["fuente"], 99)
            p_old = SOURCE_PRIORITY.get(existing["fuente"], 99)
            if p_new < p_old:
                best[key] = rec
            elif p_new == p_old:
                # merge: prefer non-empty fields from the new record
                for f in ["nombre_modulo", "carrera", "jornada"]:
                    if not existing[f] and rec[f]:
                        existing[f] = rec[f]

    result = sorted(best.values(), key=lambda r: (r["semestre"], r["codigo_asignatura"]))
    result += sorted(uncoded, key=lambda r: (r["semestre"], r["nombre_modulo"]))
    return result


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("📂 Leyendo fuentes en:", ROOT)

    print("  [1/3] Escaneando insumos xlsx...")
    recs_ins  = collect_from_insumos(ROOT)
    print(f"        → {len(recs_ins)} registros")

    print("  [2/3] Escaneando carpetas de guías...")
    recs_gui  = collect_from_guias(ROOT)
    print(f"        → {len(recs_gui)} registros")

    print("  [3/3] Escaneando planificación xlsx...")
    recs_plan = collect_from_planificacion(ROOT)
    print(f"        → {len(recs_plan)} registros")

    all_records = recs_ins + recs_gui + recs_plan
    print(f"\n  Total antes de deduplicar: {len(all_records)}")

    final = deduplicate(all_records)
    print(f"  Total tras deduplicar:     {len(final)}")

    OUTPUT_CSV.parent.mkdir(exist_ok=True)
    df = pd.DataFrame(final, columns=COLS)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\n✅ Catálogo generado: {OUTPUT_CSV}")

    # validation preview
    print("\n── Primeras 10 filas ────────────────────────────────────────────────")
    print(df.head(10).to_string(index=False))

    print("\n── Estadísticas ─────────────────────────────────────────────────────")
    print(f"  Con código:      {df['codigo_asignatura'].ne('').sum()}")
    print(f"  Sin código:      {df['codigo_asignatura'].eq('').sum()}")
    print(f"  Por fuente:\n{df['fuente'].value_counts().to_string()}")
    print(f"  Por carrera:\n{df['carrera'].value_counts().to_string()}")
    print(f"  Por semestre:\n{df['semestre'].value_counts().to_string()}")
    print("\n✅ Logs en catalogo.log")


if __name__ == "__main__":
    main()
